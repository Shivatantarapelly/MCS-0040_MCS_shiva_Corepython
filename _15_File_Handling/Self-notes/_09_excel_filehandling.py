'''
we can interact with excel using python but by using some modules
Openpyxl is a Python library that provides various methods to interact with Excel Files using Python.
It allows operations like reading, writing, arithmetic operations, plotting graphs, etc.
'''

import openpyxl

wbook = openpyxl.load_workbook('file1.xlsx')
sheet = wbook.sheetnames
print(sheet)
print(wbook.active.title)
sh1 = wbook['data']
print(sh1['A2'].value)
print(wbook['data']['A2'].value)  # accessing data present in the sheet

row = sh1.max_row
column = sh1.max_column
print(f'maximum rows:{row}, columns :{column}')  # get max row and columns count

for i in range(1, row + 1):
    for j in range(1, column + 1):
        print(f"{sh1.cell(i, j).value}")  # reading all data in the sheet

# c = sheet['A5']
# c.value = 'shiva prasad'
# c = sheet['B5']
# c.value = 26
# c = sheet['C5']
# c.value = 'mumbai'
# wbook.save('file1.xlsx')
